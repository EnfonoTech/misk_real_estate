"""
One-time import script: reads Building Details.xlsx and creates
Item Groups (buildings), Items (units), and Item Prices in ERPNext.

Also seeds master data: Floor, Unit Type, Payment Plan, Price List records.

Usage:
    bench --site <site> execute misk_real_estate.utils.import_units.run
"""

import frappe
from frappe.utils import flt


# ── Column mappings per sheet ─────────────────────────────────────────────────

# Excel rows have a leading None at index 0 (empty column A).
# All indices are direct 0-based positions in the row tuple.

# Misk Walk: Unit No | Type | Area | Min DP | Floor | 20% price | 50% price | 100% price
MISK_WALK_COLS = {
    "unit_no": 1, "unit_type": 2, "area": 3, "min_dp": 4,
    "floor": 5, "price_20": 6, "price_50": 7, "price_100": 8,
}

# Standard sheets: Floor | Unit No | Type | Selling Price
STANDARD_COLS = {
    "floor": 1, "unit_no": 2, "unit_type": 3, "price_100": 4,
}

BUILDINGS = {
    "Misk Walk":          {"sheet": "Misk Walk",    "cols": "misk_walk"},
    "Azz":                {"sheet": "Azz",           "cols": "standard"},
    "Souq Misk":          {"sheet": "Souq Misk",     "cols": "standard"},
    "Reef":               {"sheet": "Reef",           "cols": "standard"},
    "Misk Al Mawalah":    {"sheet": "Misk Mawala",   "cols": "standard"},
    "Furath":             {"sheet": "Furath",         "cols": "standard"},
}

PRICE_LISTS = {
    "Full Payment":  "price_100",
    "50% DP Plan":   "price_50",
    "20% DP Plan":   "price_20",
}

# ── Seed master data ──────────────────────────────────────────────────────────

def _seed_masters():
    floors = ["GROUND", "FIRST", "SECOND", "THIRD", "FOURTH", "FIFTH", "ROOF", "MEZZANINE"]
    for f in floors:
        if not frappe.db.exists("Floor", f):
            frappe.get_doc({"doctype": "Floor", "floor_name": f}).insert(ignore_permissions=True)

    unit_types = ["SHOP", "1 BHK", "2 BHK", "STUDIO", "3 BHK"]
    for t in unit_types:
        if not frappe.db.exists("Unit Type", t):
            frappe.get_doc({"doctype": "Unit Type", "type_name": t}).insert(ignore_permissions=True)

    plans = [
        {"plan_name": "Full Payment",    "number_of_installments": 0,  "is_full_payment": 1},
        {"plan_name": "Installment 12M", "number_of_installments": 12, "is_full_payment": 0},
        {"plan_name": "Installment 24M", "number_of_installments": 24, "is_full_payment": 0},
        {"plan_name": "Installment 36M", "number_of_installments": 36, "is_full_payment": 0},
    ]
    for p in plans:
        if not frappe.db.exists("Payment Plan", p["plan_name"]):
            frappe.get_doc({"doctype": "Payment Plan", **p}).insert(ignore_permissions=True)

    for pl_name in PRICE_LISTS:
        if not frappe.db.exists("Price List", pl_name):
            frappe.get_doc({
                "doctype": "Price List",
                "price_list_name": pl_name,
                "currency": "OMR",
                "buying": 0,
                "selling": 1,
                "enabled": 1,
            }).insert(ignore_permissions=True)

    frappe.db.commit()
    frappe.logger().info("import_units: masters seeded")


# ── Ensure Item Group (building) exists ───────────────────────────────────────

def _ensure_item_group(building_name):
    if frappe.db.exists("Item Group", building_name):
        return
    frappe.get_doc({
        "doctype": "Item Group",
        "item_group_name": building_name,
        "parent_item_group": "All Item Groups",
        "is_group": 0,
    }).insert(ignore_permissions=True)


# ── Create or update a single unit Item ──────────────────────────────────────

def _ensure_unit_type(type_name):
    if type_name and not frappe.db.exists("Unit Type", type_name):
        frappe.get_doc({"doctype": "Unit Type", "type_name": type_name}).insert(ignore_permissions=True)


def _ensure_floor(floor_name):
    if floor_name and not frappe.db.exists("Floor", floor_name):
        frappe.get_doc({"doctype": "Floor", "floor_name": floor_name}).insert(ignore_permissions=True)


def _upsert_item(item_code, item_name, building, unit_type, floor, area, min_dp):
    _ensure_unit_type(unit_type)
    _ensure_floor(floor)

    if frappe.db.exists("Item", item_code):
        # Update custom fields only
        frappe.db.set_value("Item", item_code, {
            "unit_type": unit_type or "",
            "floor_number": floor or "",
            "unit_area_sqft": flt(area),
            "min_down_payment_pct": flt(min_dp) * 100 if flt(min_dp) <= 1 else flt(min_dp),
        })
        return item_code

    item = frappe.get_doc({
        "doctype": "Item",
        "item_code": item_code,
        "item_name": item_name,
        "item_group": building,
        "is_sales_item": 1,
        "is_purchase_item": 0,
        "is_stock_item": 0,
        "unit_status": "Available",
        "unit_type": unit_type or "",
        "floor_number": floor or "",
        "unit_area_sqft": flt(area),
        "min_down_payment_pct": flt(min_dp) * 100 if flt(min_dp) <= 1 else flt(min_dp),
    })
    item.insert(ignore_permissions=True)
    return item_code


# ── Set Item Price ─────────────────────────────────────────────────────────────

def _set_item_price(item_code, price_list, price):
    if not price or flt(price) <= 0:
        return
    existing = frappe.db.get_value("Item Price", {"item_code": item_code, "price_list": price_list}, "name")
    if existing:
        frappe.db.set_value("Item Price", existing, "price_list_rate", flt(price))
    else:
        frappe.get_doc({
            "doctype": "Item Price",
            "item_code": item_code,
            "price_list": price_list,
            "price_list_rate": flt(price),
            "currency": "OMR",
        }).insert(ignore_permissions=True)


# ── Parse one sheet ───────────────────────────────────────────────────────────

def _parse_sheet(ws, building_name, col_mode):
    created = 0
    for row in ws.iter_rows(min_row=5, values_only=True):  # skip header rows
        if col_mode == "misk_walk":
            unit_no   = row[MISK_WALK_COLS["unit_no"]]
            unit_type = row[MISK_WALK_COLS["unit_type"]]
            area      = row[MISK_WALK_COLS["area"]]
            min_dp    = row[MISK_WALK_COLS["min_dp"]]
            floor     = row[MISK_WALK_COLS["floor"]]
            price_100 = row[MISK_WALK_COLS["price_100"]]
            price_50  = row[MISK_WALK_COLS["price_50"]]
            price_20  = row[MISK_WALK_COLS["price_20"]]
        else:
            floor     = row[STANDARD_COLS["floor"]]
            unit_no   = row[STANDARD_COLS["unit_no"]]
            unit_type = row[STANDARD_COLS["unit_type"]]
            price_100 = row[STANDARD_COLS["price_100"]]
            area = min_dp = price_50 = price_20 = None

        if not unit_no:
            continue

        # Normalise floor name to uppercase
        floor_str = str(floor).strip().upper() if floor else ""

        # Build item_code: "BUILDING PREFIX - UNIT NO"
        prefix = building_name.upper().replace(" ", "-")[:10]
        item_code = f"{prefix}-{str(unit_no).strip()}"
        item_name = str(unit_no).strip()

        _upsert_item(item_code, item_name, building_name,
                     str(unit_type).strip() if unit_type else "",
                     floor_str, area, min_dp)

        _set_item_price(item_code, "Full Payment", price_100)
        if price_50:
            _set_item_price(item_code, "50% DP Plan", price_50)
        if price_20:
            _set_item_price(item_code, "20% DP Plan", price_20)

        created += 1

    return created


# ── Main entry point ──────────────────────────────────────────────────────────

def run():
    import openpyxl
    import os

    xlsx_path = os.path.join(frappe.get_site_path(), "..", "..", "Building Details.xlsx")
    if not os.path.exists(xlsx_path):
        # Try bench root
        xlsx_path = "/home/ramees/frappe-bench/Building Details.xlsx"

    if not os.path.exists(xlsx_path):
        frappe.throw(f"Building Details.xlsx not found at {xlsx_path}")

    frappe.logger().info("import_units: seeding masters...")
    _seed_masters()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    total = 0
    for building_name, cfg in BUILDINGS.items():
        sheet_name = cfg["sheet"]
        if sheet_name not in wb.sheetnames:
            frappe.logger().warning(f"import_units: sheet '{sheet_name}' not found, skipping")
            continue

        _ensure_item_group(building_name)
        ws = wb[sheet_name]
        count = _parse_sheet(ws, building_name, cfg["cols"])
        frappe.db.commit()
        frappe.logger().info(f"import_units: {building_name} — {count} units processed")
        total += count

    frappe.logger().info(f"import_units: done. Total units processed: {total}")
    print(f"Done. {total} units imported.")


@frappe.whitelist()
def get_import_template():
    """
    Download a single-sheet Building Details template.
    Columns: Floor No. | Unit No. | Type | Area (sqm) | <Price List columns>
    Price list column headers must match an existing Price List name exactly.
    """
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Building Name"

    header_fill = PatternFill("solid", fgColor="1F3864")
    price_fill  = PatternFill("solid", fgColor="2E75B6")
    note_fill   = PatternFill("solid", fgColor="FFF2CC")
    hdr_font    = Font(color="FFFFFF", bold=True)

    # Row 1: building name instruction
    ws["B1"] = "BUILDING NAME : YOUR BUILDING NAME"
    ws["B1"].font = Font(bold=True, size=12)

    # Row 3: tip about price list columns
    ws["B3"] = "Rename this sheet to your building name. Price list columns: header must exactly match a Price List name in the system (e.g. Full Payment, 50% DP Plan)."
    ws["B3"].font = Font(color="7F6000", italic=True)
    ws["B3"].fill = note_fill
    ws.merge_cells("B3:H3")

    # Row 4: headers — fixed columns + sample price lists
    headers = [None, "Floor No.", "Unit No.", "Type", "Area (sqm)", "Full Payment", "50% DP Plan", "20% DP Plan"]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=val)
        if val:
            cell.font = hdr_font
            cell.fill = header_fill if col <= 5 else price_fill
            cell.alignment = Alignment(horizontal="center")

    for col, width in [(2,12),(3,12),(4,12),(5,12),(6,16),(7,14),(8,14)]:
        ws.column_dimensions[get_column_letter(col)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.local.response.filename = "Building Details Template.xlsx"
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "download"


@frappe.whitelist()
def run_from_file(file_url):
    """
    Called from the Item list view "Import Units from Excel" button.
    Reads the uploaded .xlsx file and imports all sheets as buildings.
    Price list columns are detected dynamically from the header row —
    any column whose header matches an existing Price List name is imported.
    """
    import openpyxl
    from io import BytesIO

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    content = file_doc.get_content()

    frappe.logger().info("import_units: seeding masters...")
    _seed_masters()

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

    # Get all Price List names for dynamic column detection
    known_price_lists = set(frappe.db.get_all("Price List", pluck="name"))

    total = 0
    processed_buildings = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        building_name = _get_building_name_from_sheet(ws, sheet_name)
        _ensure_item_group(building_name)
        count = _parse_sheet_dynamic(ws, building_name, known_price_lists)
        frappe.db.commit()
        frappe.logger().info(f"import_units: {building_name} — {count} units processed")
        total += count
        processed_buildings.append(building_name)

    msg = f"Import complete — {total} units across {len(processed_buildings)} buildings: {', '.join(processed_buildings)}"
    frappe.logger().info(f"import_units: {msg}")
    return msg


def _get_building_name_from_sheet(ws, sheet_name):
    """Read 'BUILDING NAME : ...' from row 2, fallback to sheet name."""
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "BUILDING NAME" in cell.upper():
                parts = cell.split(":")
                if len(parts) > 1:
                    return parts[1].strip()
    return sheet_name


def _parse_sheet_dynamic(ws, building_name, known_price_lists):
    """
    Dynamic sheet parser: detects column positions from header row.
    Required columns: Unit No. (or Unit No), Type (or unit_type), Floor No.
    Price list columns: any header that matches a Price List name exactly.
    """
    # Find the header row (first row where 'unit no' or 'floor' appears)
    header_row_idx = None
    headers = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        row_lower = [str(c).strip().lower() if c else "" for c in row]
        if any(k in " ".join(row_lower) for k in ["unit no", "floor no", "unit no."]):
            header_row_idx = row_idx
            for col_idx, val in enumerate(row):
                if val:
                    headers[str(val).strip()] = col_idx
            break

    if not header_row_idx:
        frappe.logger().warning(f"import_units: no header row found in sheet '{building_name}', skipping")
        return 0

    # Map column indices for required fields (case-insensitive)
    def find_col(keywords):
        for k, idx in headers.items():
            if any(kw.lower() in k.lower() for kw in keywords):
                return idx
        return None

    col_unit_no   = find_col(["unit no"])
    col_unit_type = find_col(["type"])
    col_floor     = find_col(["floor"])
    col_area      = find_col(["area"])
    col_min_dp    = find_col(["min dp", "min down"])

    # Detect price list columns
    price_cols = {}  # {price_list_name: col_idx}
    for header, col_idx in headers.items():
        if header in known_price_lists:
            price_cols[header] = col_idx

    if col_unit_no is None:
        frappe.logger().warning(f"import_units: 'Unit No.' column not found in '{building_name}', skipping")
        return 0

    created = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        unit_no = row[col_unit_no] if col_unit_no is not None else None
        if not unit_no:
            continue

        unit_type = str(row[col_unit_type]).strip() if col_unit_type is not None and row[col_unit_type] else ""
        floor_val = str(row[col_floor]).strip().upper() if col_floor is not None and row[col_floor] else ""
        area      = row[col_area] if col_area is not None else None
        min_dp    = row[col_min_dp] if col_min_dp is not None else None

        prefix = building_name.upper().replace(" ", "-")[:10]
        item_code = f"{prefix}-{str(unit_no).strip()}"
        item_name = str(unit_no).strip()

        _upsert_item(item_code, item_name, building_name, unit_type, floor_val, area, min_dp)

        for pl_name, col_idx in price_cols.items():
            price = row[col_idx] if col_idx < len(row) else None
            if price:
                _set_item_price(item_code, pl_name, price)

        created += 1

    return created
